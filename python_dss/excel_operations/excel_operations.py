"""
Операции с Excel файлами (аналог EPPlus)
"""

from typing import Optional, Union
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelOperations:
    """Класс для работы с Excel файлами"""
    
    def __init__(self, file: Optional[str] = None, sheet_name: str = "Результат"):
        self._workbook = None
        self._worksheet = None
        
        if file and Path(file).exists():
            self._workbook = load_workbook(file)
            if sheet_name in self._workbook.sheetnames:
                self._worksheet = self._workbook[sheet_name]
            else:
                self._worksheet = self._workbook.active
        else:
            self._workbook = Workbook()
            self._worksheet = self._workbook.active
            self._worksheet.title = sheet_name
        
        # Установка переноса текста для всех ячеек
        self._set_wrap_text_all()
    
    def _set_wrap_text_all(self):
        """Установка переноса текста для всех ячеек"""
        for row in self._worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
    
    def add_list(self, sheet_name: str):
        """Добавление нового листа"""
        self._worksheet = self._workbook.create_sheet(sheet_name)
        self._set_wrap_text_all()
    
    def set_val(self, row: int, col: int, value: Union[str, int, float], wrap: bool = True):
        """Установка значения в ячейку"""
        cell = self._worksheet.cell(row, col)
        cell.value = value
        if wrap:
            cell.alignment = Alignment(wrap_text=True)
    
    def set_val_by_name(self, cell_name: str, value: Union[str, int, float]):
        """Установка значения по имени ячейки (например, 'A1')"""
        self._worksheet[cell_name] = value
    
    def get_str(self, row: int, col: int) -> str:
        """Получение строкового значения"""
        cell = self._worksheet.cell(row, col)
        return str(cell.value) if cell.value is not None else ""
    
    def get_int(self, row: int, col: int) -> int:
        """Получение целочисленного значения"""
        cell = self._worksheet.cell(row, col)
        try:
            return int(cell.value) if cell.value is not None else 0
        except:
            return 0
    
    def get_dbl(self, row: int, col: int) -> float:
        """Получение вещественного значения"""
        cell = self._worksheet.cell(row, col)
        try:
            return float(cell.value) if cell.value is not None else 0.0
        except:
            return 0.0
    
    def save(self, file: Optional[str] = None):
        """Сохранение файла"""
        if file:
            self._workbook.save(file)
        else:
            # Сохранение во временный файл
            tmp_file = Path.cwd() / "tmp.xlsx"
            self._workbook.save(tmp_file)
    
    def borders(self, start_row: int, start_col: int, end_row: int, end_col: int):
        """Установка границ для диапазона ячеек"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                self._worksheet.cell(row, col).border = thin_border
    
    def format_cells(self, start_row: int, start_col: int, end_row: int, end_col: int,
                     bold: bool = False, italic: bool = False, color: Optional[int] = None):
        """Форматирование ячеек"""
        font = Font(bold=bold, italic=italic)
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = self._worksheet.cell(row, col)
                cell.font = font
                
                if color is not None:
                    fill = PatternFill(start_color=f"{color:06X}", end_color=f"{color:06X}", fill_type="solid")
                    cell.fill = fill
    
    def merge(self, start_row: int, start_col: int, end_row: int, end_col: int,
              horizontal: bool = False, vertical: bool = False):
        """Объединение ячеек"""
        start_cell = get_column_letter(start_col) + str(start_row)
        end_cell = get_column_letter(end_col) + str(end_row)
        
        self._worksheet.merge_cells(f"{start_cell}:{end_cell}")
        
        if horizontal or vertical:
            alignment = Alignment(
                horizontal='center' if horizontal else 'general',
                vertical='center' if vertical else 'bottom'
            )
            self._worksheet[start_cell].alignment = alignment
    
    def format(self, start_row: int, start_col: int, end_row: int, end_col: int,
               horizontal: str = 'general', vertical: str = 'bottom', rotation: int = 0):
        """Форматирование выравнивания"""
        alignment = Alignment(
            horizontal=horizontal,
            vertical=vertical,
            text_rotation=rotation
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                self._worksheet.cell(row, col).alignment = alignment
    
    def font(self, name: str = "Times New Roman", size: int = 10):
        """Установка шрифта для всех ячеек"""
        font = Font(name=name, size=size)
        
        for row in self._worksheet.iter_rows():
            for cell in row:
                cell.font = font
    
    def width(self, col: int, width: float, auto: bool = False):
        """Установка ширины столбца"""
        col_letter = get_column_letter(col)
        if auto:
            self._worksheet.column_dimensions[col_letter].width = None  # Автоподбор
        else:
            self._worksheet.column_dimensions[col_letter].width = width
    
    def hide_column(self, col: int):
        """Скрытие столбца"""
        col_letter = get_column_letter(col)
        self._worksheet.column_dimensions[col_letter].hidden = True
    
    def height(self, row: int, height: float):
        """Установка высоты строки"""
        self._worksheet.row_dimensions[row].height = height
    
    def val_to_color(self, value: float) -> int:
        """Преобразование значения в цвет (для цветовой индикации)"""
        # Цветовая схема по значениям
        if value >= 30 and value < 40:
            return 0x9ACD32  # YellowGreen
        elif value >= 40 and value < 50:
            return 0x90EE90  # LightGreen
        elif value >= 50 and value < 60:
            return 0xADFF2F  # GreenYellow
        elif value >= 60 and value < 70:
            return 0xFFFF00  # Yellow
        elif value >= 70 and value < 80:
            return 0xFFA500  # Orange
        elif value >= 80 and value < 90:
            return 0xF4A460  # SandyBrown
        elif value >= 90 and value < 100:
            return 0xFF6347  # Tomato
        elif value >= 100:
            return 0xFF4500  # OrangeRed
        else:
            return 0xFFFFFF  # White
    
    def val_to_color_voltage(self, value: float) -> int:
        """Преобразование напряжения в цвет"""
        if 10 <= value <= 15:
            return 0xADFF2F  # GreenYellow
        elif 8 <= value < 10:
            return 0xFFFF00  # Yellow
        elif 6 <= value < 8:
            return 0xFFA500  # Orange
        elif 4 <= value < 6:
            return 0xF4A460  # SandyBrown
        elif 2.5 <= value < 4:
            return 0xFF6347  # Tomato
        elif value <= 2.5:
            return 0xFF4500  # OrangeRed
        else:
            return 0xFFFFFF  # White
    
    def voltage_to_color(self, value: float) -> int:
        """Преобразование напряжения в цвет (альтернативная схема)"""
        if value >= 16:
            return 0xFF4500  # OrangeRed
        elif 14 <= value < 16:
            return 0xFF6347  # Tomato
        elif 12 <= value < 14:
            return 0xF4A460  # SandyBrown
        elif 10 <= value < 12:
            return 0xFFA500  # Orange
        elif 7.5 <= value < 10:
            return 0xFFFF00  # Yellow
        elif 5 <= value < 7.5:
            return 0xADFF2F  # GreenYellow
        else:
            return 0xFFFFFF  # White

